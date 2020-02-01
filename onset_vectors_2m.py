#!/usr/bin/env python
# coding: utf-8

# Study-specific script for creating onset vector files. Needs to be modified for other conditions.
# Columns starting from A1: Group, Subject, Trial, accuracy, condition number, onset time
# Data sorted smallest to largest in order: Group, Subject, accuracy, condition number, Trial 

import os
import openpyxl

wb = openpyxl.load_workbook('onset_vectors.xlsx')
sheet = wb['Sheet1']

subs = [sheet['B'+str(i)].value for i in range(2,sheet.max_row+1)]
subs = sorted(list(set(subs)))
conds = [sheet.cell(row=1,column=i).value for i in range(8,17)] 

# BUG. All rows are written to the None file each loop. Script works but is super slow because of this.
 
for s in range(len(subs)):
 for i in range(2,sheet.max_row+1):
     if sheet['D'+str(i)].value == 0 and sheet['B'+str(i)].value == subs[s]:
         file = open(str(subs[s]) + "_Err.txt", "a")
     elif sheet['D'+str(i)].value == 1 and sheet['E'+str(i)].value == 1 and sheet['B'+str(i)].value == subs[s]:
         file = open(str(subs[s]) + "_PNSNU.txt", "a")
     elif sheet['D'+str(i)].value == 1 and sheet['E'+str(i)].value == 2 and sheet['B'+str(i)].value == subs[s]:
         file = open(str(subs[s]) + "_PNSU.txt", "a")
     elif sheet['D'+str(i)].value == 1 and sheet['E'+str(i)].value == 3 and sheet['B'+str(i)].value == subs[s]:
         file = open(str(subs[s]) + "_PSNU.txt", "a")
     elif sheet['D'+str(i)].value == 1 and sheet['E'+str(i)].value == 4 and sheet['B'+str(i)].value == subs[s]:
         file = open(str(subs[s]) + "_PSU.txt", "a")
     elif sheet['D'+str(i)].value == 1 and sheet['E'+str(i)].value == 5 and sheet['B'+str(i)].value == subs[s]:
         file = open(str(subs[s]) + "_UPNSNU.txt", "a")
     elif sheet['D'+str(i)].value == 1 and sheet['E'+str(i)].value == 6 and sheet['B'+str(i)].value == subs[s]:
         file = open(str(subs[s]) + "_UPNSU.txt", "a")
     elif sheet['D'+str(i)].value == 1 and sheet['E'+str(i)].value == 7 and sheet['B'+str(i)].value == subs[s]:
         file = open(str(subs[s]) + "_UPSNU.txt", "a")
     elif sheet['D'+str(i)].value == 1 and sheet['E'+str(i)].value == 8 and sheet['B'+str(i)].value == subs[s]:
         file = open(str(subs[s]) + "_UPSU.txt", "a")
     else:
         file = open(str(subs[s]) + "_None.txt", "a")
         
    
     file.write(str(sheet['F'+str(i)].value) + "\t" + str(3) + "\t" + str(1) + "\n")
file.close()

